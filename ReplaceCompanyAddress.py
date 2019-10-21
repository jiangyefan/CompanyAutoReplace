#!/usr/bin/env python3

from openpyxl import load_workbook
import cpca
import re
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import json



class ReplaceCompanyAddress:
    def __init__(self,filepath,parameter,sendAddress,recAddress):
        self.filepath = filepath
        self.parameter = parameter
        self.sendAddress = sendAddress
        self.recAddress = recAddress

    # 先判断有没有门店一列
    def setUp(self):
        wb = load_workbook(self.filepath)
        ws = wb.get_sheet_by_name(self.parameter)
        addressInfo=ws.cell(row=1,column=ws.max_column).value
        try:
            if re.search("门店",addressInfo).group():
                return 1
        except:
            return 0



    # 处理门店数据
    def addressInfoProcess(self):
        wb = load_workbook(self.filepath)
        ws = wb.get_sheet_by_name(self.parameter)

        #单元格样式
        blue_fill = PatternFill("solid", fgColor="305496")
        red_fill = PatternFill("solid", fgColor="ff0000")
        font = Font(color='ffffff')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            diagonal=Side(style='thin', color='000000'),
            diagonal_direction=0,
            outline=Side(style='thin', color='000000'),
            vertical=Side(style='thin', color='000000'),
            horizontal=Side(style='thin', color='000000'))
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        #添加门店title
        ws[ws.cell(row=1,column=ws.max_column+1).coordinate]="门店"
        ws[ws.cell(row=1, column=ws.max_column).coordinate].fill=blue_fill
        ws[ws.cell(row=1, column=ws.max_column).coordinate].font=font
        ws[ws.cell(row=1, column=ws.max_column).coordinate].border=border
        ws[ws.cell(row=1, column=ws.max_column).coordinate].alignment=align





        #提取所有行
        rows=[]
        for row in ws.iter_rows():
            rows.append(row)

        #判断寄件地址和收件地址的列
        for addressCoordinate in range(len(rows[0])):
            if self.sendAddress == rows[0][addressCoordinate].value:
                sendAddressCol = rows[0][addressCoordinate].column
            if self.recAddress == rows[0][addressCoordinate].value:
                recAddressCol = rows[0][addressCoordinate].column

        # 地址转换门店
        # 先排除本公司的地址
        ruleAddressAll = ["缦图", "外包", "和达高科创新服务中心", "科技园路65号", "科技园65号"]
        with open("./Administrative-divisions-of-China", "r") as f:
            load_address=json.load(f)


        for addressInfo in range(1,len(rows)):
            excelAddressInfo=(rows[addressInfo][sendAddressCol - 1].value).replace(' ','')
            for ruleAddress in ruleAddressAll:
                if ruleAddress in rows[addressInfo][sendAddressCol-1].value :
                    rows[addressInfo][ws.max_column - 1].value="行政部"
                    rows[addressInfo][ws.max_column-1].alignment=align
                    rows[addressInfo][ws.max_column-1].border=border
                    break
            else:
                try:
                    addressWord=rows[addressInfo][sendAddressCol - 1].value.split()
                    province=cpca.transform(addressWord, cut=False).省.values[0]
                    city=cpca.transform(addressWord, cut=False).市.values[0]
                    area=cpca.transform(addressWord, cut=False).区.values[0]
                    if province==city and area !="":
                        for realaddress in load_address[city][area]:
                            if "|" in realaddress["address"]:
                                for addressPara in realaddress["address"].split("|"):
                                    if addressPara in excelAddressInfo:
                                        rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                        rows[addressInfo][ws.max_column-1].alignment = align
                                        rows[addressInfo][ws.max_column-1].border = border

                                        break
                                else:
                                    continue
                                break
                            else:
                                if realaddress["address"] in excelAddressInfo:
                                    rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                    rows[addressInfo][ws.max_column-1].alignment = align
                                    rows[addressInfo][ws.max_column-1].border = border

                                    break
                        else:
                            rows[addressInfo][ws.max_column - 1].value = "行政部"
                            rows[addressInfo][ws.max_column-1].alignment = align
                            rows[addressInfo][ws.max_column-1].border = border



                    elif province!=city and area !="":
                        for realaddress in load_address[province][city][area]:
                            if "|" in realaddress["address"]:
                                for addressPara in realaddress["address"].split("|"):
                                    if addressPara in excelAddressInfo:
                                        rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                        rows[addressInfo][ws.max_column-1].alignment = align
                                        rows[addressInfo][ws.max_column-1].border = border
                                        break
                                else:
                                    continue
                                break
                            else:
                                if realaddress["address"] in excelAddressInfo:
                                    rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                    rows[addressInfo][ws.max_column-1].alignment = align
                                    rows[addressInfo][ws.max_column-1].border = border

                                    break
                        else:
                            rows[addressInfo][ws.max_column - 1].value = "行政部"
                            rows[addressInfo][ws.max_column-1].alignment=align
                            rows[addressInfo][ws.max_column-1].border=border


                    else:
                        rows[addressInfo][ws.max_column - 1].value = "地址异常,请检查地址中的市和区是否遗漏或错填"
                        rows[addressInfo][ws.max_column-1].alignment = align
                        rows[addressInfo][ws.max_column-1].fill = red_fill
                        rows[addressInfo][ws.max_column-1].border = border


                except:
                    rows[addressInfo][ws.max_column - 1].value = "寄件/收件地址含有省份字样需手工匹配"
                    rows[addressInfo][ws.max_column-1].alignment = align
                    rows[addressInfo][ws.max_column-1].fill = red_fill
                    rows[addressInfo][ws.max_column-1].border = border

                    continue




        for addressInfo in range(1, len(rows)):
            excelAddressInfo=rows[addressInfo][recAddressCol - 1].value.replace(' ','')
            if rows[addressInfo][ws.max_column - 1].value == "行政部":
                try:
                    addressWord = rows[addressInfo][recAddressCol - 1].value.split()
                    province = cpca.transform(addressWord, cut=False).省.values[0]
                    city = cpca.transform(addressWord, cut=False).市.values[0]
                    area = cpca.transform(addressWord, cut=False).区.values[0]
                    if province == city and area != "":
                        for realaddress in load_address[city][area]:
                            if "|" in realaddress["address"]:
                                for addressPara in realaddress["address"].split("|"):
                                    if addressPara in excelAddressInfo:
                                        rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                        rows[addressInfo][ws.max_column-1].alignment = align
                                        rows[addressInfo][ws.max_column-1].border = border

                                        break
                                else:
                                    continue
                                break
                            else:
                                if realaddress["address"] in excelAddressInfo:
                                    rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                    rows[addressInfo][ws.max_column-1].alignment = align
                                    rows[addressInfo][ws.max_column-1].border = border
                                    break



                    elif province != city and area != "":
                        for realaddress in load_address[province][city][area]:
                            if "|" in realaddress["address"]:
                                for addressPara in realaddress["address"].split("|"):
                                    if addressPara in excelAddressInfo:
                                        rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                        rows[addressInfo][ws.max_column-1].alignment = align
                                        rows[addressInfo][ws.max_column-1].border = border
                                        break
                                else:
                                    continue
                                break
                            else:
                                if realaddress["address"] in excelAddressInfo:
                                    rows[addressInfo][ws.max_column - 1].value = realaddress["name"]
                                    rows[addressInfo][ws.max_column-1].alignment = align
                                    rows[addressInfo][ws.max_column-1].border = border
                                    break



                    else:
                        rows[addressInfo][ws.max_column - 1].value = "地址异常,请检查地址中的市和区是否遗漏或错填"
                        rows[addressInfo][ws.max_column-1].fill = red_fill
                        rows[addressInfo][ws.max_column-1].alignment = align
                        rows[addressInfo][ws.max_column-1].border = border

                except:
                    rows[addressInfo][ws.max_column - 1].value = "寄件/收件地址含有省份字样需手工匹配"
                    rows[addressInfo][ws.max_column-1].fill = red_fill
                    rows[addressInfo][ws.max_column-1].alignment = align
                    rows[addressInfo][ws.max_column-1].border = border
                    continue

        wb.save(testaddress)








if __name__ == '__main__':
    testaddress="/Users/jiangjiang/Desktop/8月中通供应链.xlsx"
    test=ReplaceCompanyAddress(testaddress,"Sheet2","收件人地址","收件人地址")
    if test.setUp()==1:
        print("已有门店信息,请检查！")
    else:
        test.addressInfoProcess()